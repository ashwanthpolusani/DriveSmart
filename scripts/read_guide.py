import openpyxl
import os

# Get the project root directory
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Load the workbook
wb_path = os.path.join(project_root, 'backend/data/dataset-data-guide.xlsx')
wb = openpyxl.load_workbook(wb_path)

for sheet_name in wb.sheetnames:
    print(f"\n========== SHEET: {sheet_name} ==========")
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        print(row)
